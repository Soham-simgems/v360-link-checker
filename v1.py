import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Path to your Chrome driver
chromedriver_path = "chromedriver.exe"

# Load Excel
wb = openpyxl.load_workbook("V360 check video 0507.xlsx")
ws = wb.active

# Setup Selenium
options = Options()
# options.add_argument("--headless")  # comment out if you want to see the browser
options.add_argument("--no-sandbox")
options.add_argument("--disable-gpu")

driver = webdriver.Chrome()

for row in range(2, ws.max_row + 1):
    url = ws[f'B{row}'].value
    try:
        driver.get(url)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        # try to find the specific error div
        try:
            error_div = driver.find_element(
                By.XPATH,
                "//div[contains(@class, 'title') and contains(text(), 'Asset URL is expired')]"
            )
            ws[f'C{row}'] = "Not working"
        except:
            ws[f'C{row}'] = "Working"

    except Exception as e:
        ws[f'C{row}'] = f"Error: {str(e)}"

wb.save("checked_links100.xlsx")
driver.quit()
